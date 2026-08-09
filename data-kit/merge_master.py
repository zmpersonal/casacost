#!/usr/bin/env python3
"""Consolidate the uploaded Austin pricing database with the Wave 1 provider matrix."""
import pandas as pd, re, warnings
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
warnings.filterwarnings("ignore")

UP   = "/mnt/user-data/uploads/Austin_Home_Services_Company_Pricing_Database_2026-08-08.xlsx"
MINE = "austin-provider-matrix.xlsx"
OUT  = "/mnt/user-data/outputs/trueline/data-kit/austin-provider-matrix-MASTER.xlsx"

# ---- known true duplicates (present in both datasets) ----
DUPES = {"lawnstarter", "bluescience", "abchomecommercial"}
def key(s):  # loose company key
    k = re.sub(r"\(.*?\)", "", str(s)).lower()
    k = re.sub(r"\b(services?|commercial|austin|tx|co|inc|llc)\b", "", k)
    return re.sub(r"[^a-z0-9]", "", k)

# ---- load uploaded ----
u_cat = pd.read_excel(UP, sheet_name="Service Catalog", header=0)
u_ci  = pd.read_excel(UP, sheet_name="Company Index",   header=0)

# ---- load mine ----
m_prov = pd.read_excel(MINE, sheet_name="Providers",          header=2)
m_obs  = pd.read_excel(MINE, sheet_name="Price_Observations", header=2)
m_bench= pd.read_excel(MINE, sheet_name="Aggregator_Benchmarks", header=2)

PT_NORM = {
 "Quote only":"quote_only","Exact posted price":"advertised","Published company price":"advertised",
 "Published company range":"advertised","Starting price":"advertised","Austin-specific posted range":"advertised",
 "Current promotion":"promotion","Financing / promotion":"promotion","Expired promotion":"promotion_expired",
 "Official dated price guide":"guide","Published market-average statement":"guide",
}
def tier_from_pt(pt): return "C"  # all uploaded rows are provider-sourced public pages

# ---- build unified Service_Price_Rows ----
cols = ["Origin","Company","Vertical","Service / Plan","Included / Scope","Price Low","Price High",
        "Unit","Price Display","price_type","source_tier","Frequency","Add-ons / Exclusions",
        "Active","Confidence","Source URL","Dating Note","Date"]
rows = []
for _,r in u_cat.iterrows():
    rows.append(["Uploaded", r["Company"], r["Vertical"], r["Exact Service / Plan Name"],
        r["What Is Included / Scope"], r["Price Low"], r["Price High"], r["Price Unit"], r["Price Display"],
        PT_NORM.get(str(r["Price Type"]).strip(), str(r["Price Type"])), tier_from_pt(r["Price Type"]),
        r["Frequency"], r["Add-ons / Exclusions / Caveats"], r["Active as of 2026-08-08?"],
        r["Confidence"], r["Source URL"], r["Source / Dating Note"], r["Research Date"]])

VERT_MAP = {"Pool":"Pool","Lawn":"Lawn Care","Cleaning":"House Cleaning","Pest":"Pest Control",
            "Handyman":"Handyman","Multi-trade":"Multi-Trade","Multi-trade (HVAC/air)":"HVAC"}
for _,r in m_obs.iterrows():
    if key(r["Company"]) in DUPES:  # skip rows already covered by richer uploaded data
        continue
    conf = "Medium" if "STALE" in str(r["Included (summary)"]).upper() or "2024" in str(r["Date"]) else "High"
    rows.append(["Wave1", r["Company"], VERT_MAP.get(str(r["Category"]), r["Category"]),
        r["Service (as named)"], r["Included (summary)"], None, None, r["Unit"], r["Price"],
        r["price_type"], r["tier"], None, None, "Yes", conf, r["Source URL"], "", r["Date"]])

df = pd.DataFrame(rows, columns=cols)

# ---- build unified Companies ----
ccols = ["Company","Source","Operator Type / Scale","Pricing Posture","Categories",
         "Service Rows","Public-Price Rows","Quote-Only Rows","Primary Source URL","Notable"]
def counts_for(name):
    sub = df[df["Company"] == name]
    total = len(sub)
    quote = int((sub["price_type"] == "quote_only").sum())
    return total, total - quote, quote

comp_rows = []
for _,r in u_ci.iterrows():
    src = "Both" if key(r["Company"]) in DUPES else "Uploaded"
    comp_rows.append([r["Company"], src, r["Operator Type / Scale"], r["Pricing Posture"], "",
        r["Service Rows in Dataset"], r["Public/Non-Quote Price Rows"], r["Quote-Only Rows"],
        r["Primary Source URL"], r["Why Included as Substantial"]])
for _,r in m_prov.iterrows():
    if key(r["Company"]) in DUPES: continue
    t,p,q = counts_for(r["Company"])
    comp_rows.append([r["Company"], "Wave1", f"{r['Categories']} — {r['Business model']}",
        ("Publishes price" if str(r["Posts price?"]).strip()=="Yes" else "Quote-first"),
        r["Categories"], t, p, q, r["Website"], r["Notable"]])
comp_df = pd.DataFrame(comp_rows, columns=ccols).sort_values("Company")

# ---- recompute Vertical Coverage from merged rows ----
vc = []
for v, sub in df.groupby("Vertical"):
    total = len(sub); quote = int((sub["price_type"]=="quote_only").sum())
    ncomp = sub["Company"].nunique()
    vc.append([v, ncomp, total, total-quote, quote, round(100*(total-quote)/total) if total else 0])
vc_df = pd.DataFrame(vc, columns=["Vertical","Companies","Service Rows","Public-Price Rows",
        "Quote-Only Rows","Transparency %"]).sort_values("Service Rows", ascending=False)

# ==================== WRITE FORMATTED WORKBOOK ====================
PINE="1E3D2E"; SANDL="F4F0E6"; INK="20241F"; LINE="DED6C4"
HFONT=Font(name="Arial",bold=True,size=10,color="FFFFFF"); CELL=Font(name="Arial",size=9,color=INK)
TITLE=Font(name="Arial",bold=True,size=14,color=PINE); SUB=Font(name="Arial",size=9,color="5C6157")
HFILL=PatternFill("solid",fgColor=PINE); BAND=PatternFill("solid",fgColor=SANDL)
thin=Side(style="thin",color=LINE); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
WRAP=Alignment(wrap_text=True,vertical="top")

def sheet_from_df(wb, name, title, dframe, widths, start=3):
    ws = wb.create_sheet(name)
    ws.cell(row=1,column=1,value=title).font = TITLE
    for i,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=w
    for j,h in enumerate(dframe.columns,1):
        c=ws.cell(row=start,column=j,value=h); c.font=HFONT; c.fill=HFILL; c.border=BORDER
        c.alignment=Alignment(wrap_text=True,vertical="center")
    ws.row_dimensions[start].height=28
    r=start+1
    for _,row in dframe.iterrows():
        for j,val in enumerate(row,1):
            v = "" if (pd.isna(val) if not isinstance(val,(list,dict)) else False) else val
            c=ws.cell(row=r,column=j,value=v); c.font=CELL; c.border=BORDER; c.alignment=WRAP
        if (r-start)%2==0:
            for j in range(1,len(dframe.columns)+1): ws.cell(row=r,column=j).fill=BAND
        r+=1
    ws.freeze_panes=ws.cell(row=start+1,column=1)
    ws.auto_filter.ref=f"A{start}:{get_column_letter(len(dframe.columns))}{r-1}"
    return ws

wb=openpyxl.Workbook(); ws=wb.active; ws.title="README"
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=112
ws["B2"]="Austin Home-Services Provider Matrix — MASTER (Wave 1 + uploaded database, consolidated)"; ws["B2"].font=TITLE
readme=[
 "",
 f"Consolidates two independent research passes into one deduplicated dataset:",
 f"  • Uploaded 'Austin Home Services Company Pricing Database' (2026-08-08): 43 companies, 384 service rows.",
 f"  • Wave 1 provider matrix (this project): 25 companies, 31 price rows.",
 f"  => MASTER: {comp_df.shape[0]} companies, {df.shape[0]} service/price rows across {vc_df.shape[0]} verticals.",
 "",
 "The two barely overlap (only ~3 shared companies), so they stack rather than duplicate. The uploaded set",
 "adds ~20 verticals we didn't have (junk removal, garage door, water treatment, foundation, chimney, carpet,",
 "electrical, plumbing, HVAC, roofing, drain/sewer, tree, painting, etc.). Wave 1 adds ~21 pool / cleaning /",
 "pest operators the uploaded set lacked (Ideal Pool Care, The Pool Police, Endless Blue, Bluewater, Cowboy,",
 "Austin's Maid Service, Boardwalk, CR Maids, Absolute Pest, Alta, Massey, A-Tex, Stride, and more).",
 "",
 "TABS",
 "  • Companies — one row per company, with Source (Uploaded / Wave1 / Both), operator type, pricing posture, row counts.",
 "  • Service_Price_Rows — one row per service/plan, with normalized price_type + source_tier. THE core table.",
 "  • Vertical_Coverage — companies, rows, and pricing-transparency % per vertical (recomputed across the merge).",
 "  • Aggregator_Benchmarks — directory/guide ranges (Angi etc.), tier E CONTEXT, NOT provider prices.",
 "",
 "price_type (normalized): advertised · guide · promotion · promotion_expired · membership · marketplace · quote_only",
 "source_tier: A final invoice · B written quote · C provider public page · D consumer report · E aggregator estimate",
 "",
 "HONEST STATUS (unchanged from the audit standard)",
 "  • Everything here is tier C/D/E. There are still ZERO tier-A (paid invoice) or tier-B (written-quote) observations.",
 "  • 'Quote only' dominates (the single largest price_type). That mix is itself the key market finding.",
 "  • Uploaded 'Confidence' is the author's self-assessment; a few promo rows are already flagged expired.",
 "  • These are inputs to the price engine, not publishable Austin medians. Publish only after the tier gates in",
 "    segment-targets.csv are met (25+ scope-matched obs for a range; 50+ w/ >=40% tier A/B for HIGH).",
]
r=4
for line in readme:
    c=ws.cell(row=r,column=2,value=line); c.font=Font(name="Arial",bold=True,size=9,color=INK) if (line.isupper() and line) else SUB
    c.alignment=Alignment(wrap_text=True,vertical="top"); r+=1

sheet_from_df(wb,"Companies","Companies — consolidated (Wave 1 + uploaded)",comp_df,
    [34,10,34,22,22,9,9,9,40,46])
sheet_from_df(wb,"Service_Price_Rows","Service & price rows — consolidated",df,
    [26,16,30,40,10,10,12,22,14,10,10,22,16,8,10,40,30,12])
sheet_from_df(wb,"Vertical_Coverage","Coverage & pricing transparency by vertical",vc_df,
    [24,12,13,16,14,14])
sheet_from_df(wb,"Aggregator_Benchmarks","Aggregator / guide benchmarks — tier E, NOT provider prices",
    m_bench,[16,14,30,22,34,34,12])

wb.save(OUT)
print("Saved", OUT)
print(f"Companies: {comp_df.shape[0]} | Service rows: {df.shape[0]} | Verticals: {vc_df.shape[0]}")
print(f"Quote-only rows: {int((df['price_type']=='quote_only').sum())} | Priced rows: {int((df['price_type']!='quote_only').sum())}")
