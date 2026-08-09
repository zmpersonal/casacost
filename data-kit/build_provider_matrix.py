#!/usr/bin/env python3
"""Builds the Austin home-services provider matrix (Wave 1) as a formatted .xlsx."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = "/mnt/user-data/outputs/trueline/data-kit/austin-provider-matrix.xlsx"

INK   = "20241F"; PINE = "1E3D2E"; SAND = "ECE6D8"; SANDL = "F4F0E6"; LINE = "DED6C4"
HFONT = Font(name="Arial", bold=True, size=11, color="FFFFFF")
TITLE = Font(name="Arial", bold=True, size=15, color=PINE)
SUB   = Font(name="Arial", size=10, color="5C6157")
CELL  = Font(name="Arial", size=10, color=INK)
BOLD  = Font(name="Arial", bold=True, size=10, color=INK)
HFILL = PatternFill("solid", fgColor=PINE)
BAND  = PatternFill("solid", fgColor=SANDL)
thin  = Side(style="thin", color=LINE)
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
WRAP  = Alignment(wrap_text=True, vertical="top")
TOP   = Alignment(vertical="top")

def style_header(ws, row, ncols):
    for c in range(1, ncols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = HFONT; cell.fill = HFILL; cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[row].height = 30

def write_table(ws, start, headers, rows, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for j, h in enumerate(headers, 1):
        ws.cell(row=start, column=j, value=h)
    style_header(ws, start, len(headers))
    r = start + 1
    for row in rows:
        for j, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = CELL; cell.border = BORDER; cell.alignment = WRAP
        if (r - start) % 2 == 0:
            for j in range(1, len(headers)+1):
                ws.cell(row=r, column=j).fill = BAND
        r += 1
    ws.freeze_panes = ws.cell(row=start+1, column=1)
    ws.auto_filter.ref = f"{get_column_letter(1)}{start}:{get_column_letter(len(headers))}{r-1}"
    return r

wb = openpyxl.Workbook()

# ---------------------------------------------------------------- README
ws = wb.active; ws.title = "README"
ws.column_dimensions["A"].width = 3
ws.column_dimensions["B"].width = 110
ws["B2"] = "Austin Home-Services Provider Matrix — Wave 1"; ws["B2"].font = TITLE
notes = [
 "",
 "A seed market-intelligence dataset, not a complete census. Austin has hundreds-to-thousands of relevant",
 "businesses per category; this Wave 1 captures substantial operators and every published price we could verify,",
 "across pool, lawn/landscape, cleaning, handyman, and pest, plus the multi-trade incumbents.",
 "",
 "HOW TO READ IT",
 "• Providers tab — one row per company: what they do, service area, business model, and whether they post prices.",
 "• Price_Observations tab — one row per observed price, with price_type + source_tier so nothing gets pooled wrongly.",
 "• Aggregator_Benchmarks tab — directory/guide ranges (Angi etc.). These are tier E context, NOT provider prices.",
 "• Sources tab — every URL, publisher, and the date we checked it.",
 "",
 "PRICE_TYPE (never average across types)",
 "  advertised = public list / 'starting at' price       membership = subscription / plan price",
 "  marketplace = instant/bid price via a booking platform  quote_only = no public price; inspect-then-quote",
 "",
 "SOURCE_TIER (evidence quality)",
 "  A final invoice · B written quote · C provider public price page · D consumer self-report · E aggregator estimate",
 "  Wave 1 is almost entirely tier C (provider pages) and tier E (guides). Tier A/B come from the collection kit.",
 "",
 "IMPORTANT CAVEATS",
 "• 'From $X' and 'starting at' are floors, not typical paid prices. Scope and home specifics move the real number.",
 "• Quote-only is itself a finding: many substantial operators (ABC, A-Tex, Massey, most handymen) never post a price.",
 "• A few cleaning ranges come from providers' own city 'pricing guides' — provider-produced observations, not invoices.",
 "• Marketplaces (LawnGuru, GreenPal, LawnStarter) are aggregators; their 'pros' are separate local operators.",
 "• Dates matter — one cleaning guide is from 2024 and flagged stale. Re-verify before publishing any figure.",
 "",
 "NEXT WAVES (to approach 'exhaustive')",
 "  Pool 75-150 operators · Pest via TDA license list · Lawn/landscape · Cleaning/handyman long tail ·",
 "  then HVAC/plumbing/electrical/roofing (licensing-first). Extend the same columns; keep evidence + dates.",
]
r = 4
for line in notes:
    c = ws.cell(row=r, column=2, value=line)
    c.font = BOLD if line.isupper() and line else SUB
    c.alignment = Alignment(wrap_text=True, vertical="top")
    r += 1

# ---------------------------------------------------------------- PROVIDERS
ws = wb.create_sheet("Providers")
ws["A1"] = "Providers — substantial Austin-area operators (Wave 1)"; ws["A1"].font = TITLE
prov_headers = ["Company","Categories","Service area","Business model","Posts price?","Price summary","Notable","Website","Evidence date"]
prov_widths  = [24,26,26,20,12,40,34,34,13]
providers = [
 ["Ideal Pool Care","Pool","Austin","Weekly route","Yes","Weekly service $225–$325/mo","Price scales with size/equipment/chem demand","idealpoolco.com","2026-04"],
 ["The Pool Police","Pool","Austin","Weekly route + repair","Yes","Weekly $195–$385/mo (flat, chems incl); leak detection per visit","No contracts; tiers by pool complexity","thepoolpolice.com","2026-04"],
 ["Blue Science (Austin)","Pool","Austin (res & comm)","Weekly route","Yes","Full & Partial plans from $189/mo","Licensed/certified; per-visit report","bluescience.com","2026-08"],
 ["Endless Blue Pools","Pool","Austin & Round Rock","Weekly route + repair","Yes","Weekly from $275/mo; green-to-clean $250–$600+","No contracts; new-customer $50 off","endlessbluepools.com","2026-04"],
 ["Bluewater Pools","Pool","Austin & San Antonio","Weekly route","Yes","Full $240–$320 · Partial $150–$189 · Premium $300+/mo","Publishes detailed Austin pricing guide","bluepoolwater.com","2026-04"],
 ["Hill Country Pools","Pool","Austin","Weekly route","Yes","Weekly $120–$370/mo (≤15k-gal benchmark)","Prices by pool size/volume","hillcountrypoolsaustin.com","2026-07"],
 ["Cowboy Pools","Pool","Austin only","Productized / subscription","Yes","Sub from $99 · one-time $80 · green-to-clean $250","Checkout-style productized offers","cowboypools.com","2025-11"],
 ["LawnGuru","Lawn (marketplace)","Austin metro","Booking platform","Yes","Mowing $37.93–$44.98/cut (avg $41.45)","Instant address-based pricing","lawnguru.co","2026-08"],
 ["GreenPal","Lawn (marketplace)","Austin metro","Booking platform","Yes","Mowing avg ~$39/cut (bids ~$33–$50)","Pay-after, photo-verified; pros are separate operators","yourgreenpal.com","2025-11"],
 ["LawnStarter","Lawn (marketplace)","Austin metro","Booking platform","Yes","Mowing from $19; +fert/aeration/trimming","Pros carry liability insurance","lawnstarter.com","2026-08"],
 ["GoMow","Lawn/landscape","Austin metro","Recurring route","Quote","Recurring mowing (quote-based)","Local operator since 2016","gomow.com","2025-12"],
 ["Austin's Maid Service","Cleaning","13 Central TX cities","Recurring + one-time","Yes","Recurring from $99/visit; deep/move-out by quote","Insured, background-checked; bilingual market","austinsmaidservice.com","2026-03"],
 ["The Boardwalk Cleaning Co.","Cleaning","Austin","Flat hourly","Yes","$90–$95/hr, team of two, supplies incl","Eco supplies standard; no door upsells","boardwalkcleaning.com","2026-03"],
 ["CR Maids","Cleaning","Austin","Flat by size","Yes","Biweekly $100–$220; std 3BR $130–$170; move-out","Flat rates by home size","crmaids.com","2026-06"],
 ["Sparkly Maid Austin","Cleaning","Austin","Per-visit tiers","Yes","Weekly $75–$150 / biweekly $100–$200 (by size)","NOTE: guide dated 2024 — re-verify","sparklymaidaustin.com","2024-01 (stale)"],
 ["HandyMatt","Handyman","Austin","Labor / project","Quote","Quote-based; residential & commercial","Insured; broad task range","handymattaustin.com","2025-08"],
 ["HomePoint","Handyman / home maintenance","Austin","Membership","Yes","$4,000–$6,000/yr by sqft; extra handyman $95/hr","Quarterly visits, materials, annual inspection incl","gohomepoint.com","2025-11"],
 ["Anything Around the House","Handyman","Austin","Labor / project","Quote","Quote-only","Local, 20+ years","anythingaroundthehouse.com","2025-11"],
 ["Absolute Pest Management","Pest","Central TX / Austin","Recurring quarterly","Yes","Initial $69.99 w/ plan (std $169.99); sqft-scaled","Family-owned 25+ yrs; free inspections","absolutepestmgmt.com","2025-11"],
 ["Alta Pest Control","Pest","Austin","Recurring quarterly","Quote","Quarterly premium (general+outdoor); no one-time","Free re-treats; termite (Sentricon), wildlife","altapestcontrol.com","2026-06"],
 ["Massey Services","Pest","Austin (national)","Recurring quarterly","Quote","Quarterly; free inspection; money-back guarantee","National chain","masseyservices.com","2025-07"],
 ["A-Tex Pest Management","Pest","Austin & surrounding","Recurring + one-time","Quote","Quote-only; 10% off first recurring service","Local since 2002","atexpest.com","2025-11"],
 ["Stride Pest Control","Pest","Austin","Recurring","Quote","Quote-only","States licensed/insured (TPCL)","stridepestcontrol.com","2025-11"],
 ["ABC Home & Commercial","Multi-trade","Austin metro","Full-service / quote","Quote","Quote-only across all trades","600+ licensed pros; flat-rate 'No Surprise'","abchomeandcommercial.com","2025-11"],
 ["United Home Services","Multi-trade (HVAC/air)","Austin (by ZIP)","Quote / free estimate","Quote","Quote-only; free estimate","Duct, A/C, insulation, chimney, garage door","unitedhomeservices.com","2025-11"],
]
write_table(ws, 3, prov_headers, providers, prov_widths)

# ---------------------------------------------------------------- PRICE OBSERVATIONS
ws = wb.create_sheet("Price_Observations")
ws["A1"] = "Price observations — one row per observed price (Wave 1)"; ws["A1"].font = TITLE
po_headers = ["Obs ID","Company","Category","Service (as named)","Price","Unit","price_type","tier","Included (summary)","Source URL","Date"]
po_widths  = [8,22,12,30,16,12,12,6,40,34,11]
obs = [
 ["P-001","Ideal Pool Care","Pool","Weekly Austin Pool Service","$225–$325","per month","advertised","C","Weekly clean + chemistry; repairs pre-approved","idealpoolco.com/austin-pool-service-pricing/","2026-04"],
 ["P-002","The Pool Police","Pool","Weekly service (standard pool)","$195–$265","per month","advertised","C","Flat monthly, chemicals included, no contract","thepoolpolice.com/austin-pool-service-cost/","2026-04"],
 ["P-003","The Pool Police","Pool","Weekly service (complex/spa/feature)","$285–$385","per month","advertised","C","Higher-complexity pools","thepoolpolice.com/austin-pool-service-cost/","2026-04"],
 ["P-004","Blue Science","Pool","Full Service Plan","from $189","per month","advertised","C","Clean, maintenance, chemicals & minerals","bluescience.com/swimming-pool-service/austin-tx/","2026-08"],
 ["P-005","Blue Science","Pool","Partial Service Plan","varies","per month","advertised","C","Excludes netting & vacuuming","bluescience.com/swimming-pool-service/austin-tx/","2026-08"],
 ["P-006","Endless Blue Pools","Pool","Weekly full-service","from $275","per month","advertised","C","Chemicals included; no contract","endlessbluepools.com","2026-04"],
 ["P-007","Endless Blue Pools","Pool","Green-to-clean recovery","$250–$600+","per job","advertised","C","By algae severity, size, filter condition","endlessbluepools.com","2026-04"],
 ["P-008","Bluewater Pools","Pool","Full Service","$240–$320","per month","advertised","C","Skim, vacuum, brush, chemistry, equipment check","bluepoolwater.com/blog/pool-cleaning-cost-austin/","2026-04"],
 ["P-009","Bluewater Pools","Pool","Partial Service","$150–$189","per month","advertised","C","Reduced scope","bluepoolwater.com/blog/pool-cleaning-cost-austin/","2026-04"],
 ["P-010","Bluewater Pools","Pool","One-time cleaning","$150–$300","per job","advertised","C","Single visit","bluepoolwater.com/blog/pool-cleaning-cost-austin/","2026-04"],
 ["P-011","Hill Country Pools","Pool","Weekly service (≤15k gal)","$120–$370","per month","advertised","C","Standard residential benchmark","hillcountrypoolsaustin.com/pool-cleaning-cost-in-austin-tx/","2026-07"],
 ["P-012","Cowboy Pools","Pool","Weekly/Bi-Weekly Subscription","from $99","per month","advertised","C","Austin-only productized offer","cowboypools.com/collections/service","2025-11"],
 ["P-013","Cowboy Pools","Pool","One-Time Cleaning","$80","per job","advertised","C","Single visit","cowboypools.com/collections/service","2025-11"],
 ["P-014","Cowboy Pools","Pool","Green-To-Clean","$250","per job","advertised","C","Algae recovery","cowboypools.com/collections/service","2025-11"],
 ["L-001","LawnGuru","Lawn","Lawn mowing","$37.93–$44.98","per cut","marketplace","C","Instant address-based; avg $41.45","lawnguru.co/cities/austin-tx/lawn-mowing","2026-08"],
 ["L-002","GreenPal","Lawn","Lawn mowing","~$33–$50","per cut","marketplace","C","Bid-based; avg ~$39; pay-after","yourgreenpal.com/local/lawn-care-austin-tx","2025-11"],
 ["L-003","LawnStarter","Lawn","Lawn mowing","from $19","per cut","marketplace","C","Also fert/aeration/trimming","lawnstarter.com/austin-tx","2026-08"],
 ["C-001","Austin's Maid Service","Cleaning","Recurring house cleaning","from $99","per visit","advertised","C","Weekly/biweekly/monthly; checklist","austinsmaidservice.com","2026-03"],
 ["C-002","Austin's Maid Service","Cleaning","Deep / move-in-out","custom quote","per job","quote_only","C","Priced by size/condition","austinsmaidservice.com","2026-03"],
 ["C-003","The Boardwalk Cleaning Co.","Cleaning","Standard clean (team of two)","$90–$95","per hour","advertised","C","All supplies incl; eco standard","boardwalkcleaning.com/house-cleaning-price-list-austin-tx/","2026-03"],
 ["C-004","CR Maids","Cleaning","Biweekly recurring","$100–$220","per visit","advertised","C","By home size; ~20% under one-time","crmaids.com/house-cleaning-costs/","2026-06"],
 ["C-005","CR Maids","Cleaning","Biweekly, standard 3BR","$130–$170","per visit","advertised","C","Standard 3-bed benchmark","crmaids.com/house-cleaning-costs/","2026-06"],
 ["C-006","Sparkly Maid Austin","Cleaning","Biweekly (standard home)","$100–$200","per visit","advertised","C","STALE 2024 — re-verify","sparklymaidaustin.com/blog/house-cleaning-services-austin-prices","2024-01"],
 ["H-001","HomePoint","Handyman","Home-maintenance membership (<4,999 sqft)","$4,000/yr or $1,111/qtr","per year","membership","C","4 quarterly visits, materials, 2 handyman hrs","gohomepoint.com/our-pricing/","2025-11"],
 ["H-002","HomePoint","Handyman","Extra handyman labor","$95","per hour","membership","C","Beyond included hours","gohomepoint.com/our-pricing/","2025-11"],
 ["H-003","HandyMatt","Handyman","General handyman","quote-only","per job","quote_only","C","Inspect-then-quote","handymattaustin.com","2025-08"],
 ["X-001","Absolute Pest Mgmt","Pest","Quarterly general pest — initial","$69.99 (w/ plan)","initial visit","advertised","C","$100 off std $169.99; sqft-scaled","absolutepestmgmt.com","2025-11"],
 ["X-002","Alta Pest Control","Pest","Quarterly premium plan","quote-only","per quarter","quote_only","C","General+outdoor; free re-treats; no one-time","altapestcontrol.com","2026-06"],
 ["X-003","Massey Services","Pest","Quarterly pest prevention","quote-only","per quarter","quote_only","C","Free inspection; money-back guarantee","masseyservices.com/austin/pest-control/","2025-07"],
 ["X-004","A-Tex Pest Mgmt","Pest","Residential recurring pest","quote-only","per quarter","quote_only","C","10% off first recurring service","atexpest.com","2025-11"],
 ["M-001","ABC Home & Commercial","Multi-trade","All trades","quote-only","varies","quote_only","C","Flat-rate 'No Surprise Guarantee'","abchomeandcommercial.com/austin","2025-11"],
]
write_table(ws, 3, po_headers, obs, po_widths)

# ---------------------------------------------------------------- AGGREGATOR BENCHMARKS
ws = wb.create_sheet("Aggregator_Benchmarks")
ws["A1"] = "Aggregator / guide benchmarks — tier E CONTEXT, not provider prices"; ws["A1"].font = TITLE
ab_headers = ["Source","Category","Metric","Value","Notes","URL","Date"]
ab_widths  = [16,14,30,22,34,34,11]
bench = [
 ["Angi","Pool","Pool maintenance Austin — avg","$226 ($117–$341)","Directory average","angi.com","2026-03"],
 ["Angi","Lawn","Mowing 1/4-acre — per visit","$15–$50","Avg job $116 ($48–$187)","angi.com","2026-06"],
 ["Angi","Cleaning","House cleaning — avg/visit","$169 ($120–$232)","~$55/hr; $0.09–$0.14/sqft","angi.com","2025-10"],
 ["Angi","Handyman","Handyman job — avg","$359 ($156–$565)","","angi.com","2026-06"],
 ["HomeBlue","Handyman","Hourly rate Austin","$40–$70/hr","4 hrs ~$160–$300 + materials","homeblue.com","2024"],
 ["HomeBlue","Lawn","Mowing 1/4-acre","$35–$48/cut","By lot size","homeblue.com","2026"],
 ["PestControlPricing","Pest","Quarterly plan — per visit","$100–$200","One-time avg ~$175","pestcontrolpricing.com","2026-03"],
 ["Reliant Pest","Pest","One-time treatment","$99–$299","By size/severity","reliantpest.com","2026-05"],
 ["Bluewater (guide)","Pool","TX full service","$150–$300/mo","Premium $300–$550+","bluepoolwater.com","2026-03"],
 ["Homeyou","Handyman","Avg handyman job Austin","$327–$488","Range $86–$866","homeyou.com","2026-07"],
]
write_table(ws, 3, ab_headers, bench, ab_widths)

# ---------------------------------------------------------------- SOURCES
ws = wb.create_sheet("Sources")
ws["A1"] = "Sources — every URL checked in Wave 1"; ws["A1"].font = TITLE
s_headers = ["Publisher / Company","Type","URL","Date checked"]
s_widths  = [30,22,60,13]
src = [
 ["Ideal Pool Care","Provider price page","idealpoolco.com/austin-pool-service-pricing/","2026-04"],
 ["The Pool Police","Provider price page","thepoolpolice.com/austin-pool-service-cost/","2026-04"],
 ["Blue Science","Provider page","bluescience.com/swimming-pool-service/austin-tx/","2026-08"],
 ["Endless Blue Pools","Provider price page","endlessbluepools.com/blog/pool-service-cost-austin-round-rock","2026-04"],
 ["Bluewater Pools","Provider guide/page","bluepoolwater.com/blog/pool-cleaning-cost-austin/","2026-04"],
 ["Hill Country Pools","Provider guide/page","hillcountrypoolsaustin.com/pool-cleaning-cost-in-austin-tx/","2026-07"],
 ["Cowboy Pools","Provider service page","cowboypools.com/collections/service","2025-11"],
 ["LawnGuru","Marketplace","lawnguru.co/cities/austin-tx/lawn-mowing","2026-08"],
 ["GreenPal","Marketplace","yourgreenpal.com/local/lawn-care-austin-tx","2025-11"],
 ["LawnStarter","Marketplace","lawnstarter.com/austin-tx","2026-08"],
 ["GoMow","Provider","gomow.com","2025-12"],
 ["Austin's Maid Service","Provider page","austinsmaidservice.com","2026-03"],
 ["The Boardwalk Cleaning Co.","Provider price page","boardwalkcleaning.com/house-cleaning-price-list-austin-tx/","2026-03"],
 ["CR Maids","Provider guide/page","crmaids.com/house-cleaning-costs/","2026-06"],
 ["Sparkly Maid Austin","Provider guide (stale)","sparklymaidaustin.com/blog/house-cleaning-services-austin-prices","2024-01"],
 ["HandyMatt","Provider","handymattaustin.com","2025-08"],
 ["HomePoint","Provider pricing page","gohomepoint.com/our-pricing/","2025-11"],
 ["Anything Around the House","Provider","anythingaroundthehouse.com","2025-11"],
 ["Absolute Pest Management","Provider page","absolutepestmgmt.com","2025-11"],
 ["Alta Pest Control","Provider page","altapestcontrol.com/locations/pest-control-austin-texas","2026-06"],
 ["Massey Services","Provider page","masseyservices.com/austin/pest-control/","2025-07"],
 ["A-Tex Pest Management","Provider","atexpest.com","2025-11"],
 ["Stride Pest Control","Provider","stridepestcontrol.com/austin-pest-control/","2025-11"],
 ["ABC Home & Commercial","Provider","abchomeandcommercial.com/austin","2025-11"],
 ["United Home Services","Provider","unitedhomeservices.com/locations/austin-tx/","2025-11"],
 ["Angi","Aggregator guide","angi.com (Austin cost articles)","2025-2026"],
 ["HomeBlue / Homeyou","Aggregator guide","homeblue.com; homeyou.com","2024-2026"],
 ["PestControlPricing / Reliant","Aggregator guide","pestcontrolpricing.com; reliantpest.com","2026"],
]
write_table(ws, 3, s_headers, src, s_widths)

wb.save(OUT)
print("Saved", OUT)
print("Sheets:", wb.sheetnames)
